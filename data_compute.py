import pyexcel as p
import openpyxl as xl

p.save_book_as(file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_female.xls',
               dest_file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_female.xlsx')

p.save_book_as(file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/Employment_services_female.xls',
               dest_file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/Employment_services_female.xlsx')

in_file ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_female.xlsx"
oldwb = xl.load_workbook(in_file)
oldws = oldwb.worksheets[0]

in_file1 ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/Employment_services_female.xlsx"
oldwb1 = xl.load_workbook(in_file1)
oldws1 = oldwb1.worksheets[0]

out_file = "/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_services_female.xlsx"
newwb = xl.load_workbook(out_file)
newws = newwb.active 

i=2
#read from first file
for process_line in range(2,266):
	#3,267
	country_name=oldws.cell(row = process_line, column = 1).value
	for cols in range(3,63):
		#print(oldws.cell(row =2, column = cols).value)
		data_point=oldws.cell(row = process_line, column = cols).value
		if data_point!=None:
			print(data_point)
			for line in range(2,189):
				if oldws1.cell(row = line, column = 1).value==country_name:
					data_point_2=oldws1.cell(row = line, column = cols).value
					if data_point_2!=None:
					#save data point from both file
						newws.cell(row =i, column = 1).value = country_name
						newws.cell(row =i, column = 2).value = oldws.cell(row = 1, column = cols).value
						newws.cell(row =i, column = 3).value = data_point
						newws.cell(row =i, column = 4).value = data_point_2
						i=i+1
						print("success")
		
newwb.save(str(out_file))

#employment_agriculture is done