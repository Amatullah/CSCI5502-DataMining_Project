import pyexcel as p
import openpyxl as xl

in_file ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_male.xlsx"
oldwb = xl.load_workbook(in_file)
oldws = oldwb.worksheets[0]

in_file1 ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/EmploymentToPopulationRatio_15+_male.xlsx"
oldwb1 = xl.load_workbook(in_file1)
oldws1 = oldwb1.worksheets[0]

in_file2 ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/economy_types.xlsx"
oldwb2 = xl.load_workbook(in_file2)
oldws2 = oldwb2.worksheets[0]

out_file = "/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/literacy_employment_male_6.xlsx"
newwb = xl.load_workbook(out_file)
newws = newwb.active 

i=2
#read from first file
for process_line in range(2,266):
	#3,267
	country_name=oldws.cell(row = process_line, column = 1).value
	country_code=oldws.cell(row = process_line, column = 2).value
	for p_line in range(2,189):
		country_code_economy=oldws2.cell(row = p_line, column = 1).value
		if country_code==country_code_economy:
			if oldws2.cell(row = p_line, column = 2).value=="Sub-Saharan Africa":
				for cols in range(3,63):
					#print(oldws.cell(row =2, column = cols).value)
					data_point=oldws.cell(row = process_line, column = cols).value
					if data_point!=None:
						print(data_point)
						for line in range(2,189):
							if oldws1.cell(row = line, column = 1).value==country_name:
								data_point_2=oldws1.cell(row = line, column = cols).value
								if data_point_2!=None:
								#save data point from both file
									newws.cell(row =i, column = 1).value = country_name
									newws.cell(row =i, column = 2).value = oldws.cell(row = 1, column = cols).value
									newws.cell(row =i, column = 3).value = data_point
									newws.cell(row =i, column = 4).value = data_point_2
									i=i+1
									print("success")
		
newwb.save(str(out_file))

#employment_agriculture is done
#1. East Asia & Pacific
#2. Europe & Central Asia
#3. Latin America & Caribbean
#4. Middle East & North Africa
#5. South Asia
#6. Sub-Saharan Africa



