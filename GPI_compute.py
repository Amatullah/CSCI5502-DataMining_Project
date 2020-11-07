import pyexcel as p
import openpyxl as xl

p.save_book_as(file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/GenderEmploymentIndex.xls',
               dest_file_name='/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/GenderEmploymentIndex.xlsx')


in_file ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/GenderEmploymentIndex.xlsx"
oldwb = xl.load_workbook(in_file)
oldws = oldwb.worksheets[0]

in_file1 ="/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/GPI_3_factors.xlsx"
oldwb1 = xl.load_workbook(in_file1)
oldws1 = oldwb1.worksheets[0]

out_file = "/home/xuefei/Desktop/CSCI5502-DataMining_Project/Dataset/employment/GPI_Gender_Employment_upper_middle.xlsx"
newwb = xl.load_workbook(out_file)
newws = newwb.active 

i=2
#read from first file
for process_line in range(2,189):
	#3,267
	country_name=oldws.cell(row = process_line, column = 1).value
	income_group=oldws.cell(row = process_line, column = 2).value
	if income_group=="Upper middle income":
		for cols in range(4,34):
			#print(oldws.cell(row =2, column = cols).value)
			year=oldws.cell(row = 1, column = cols).value
			data_point=oldws.cell(row = process_line, column = cols).value
			data_point_2=oldws1.cell(row = process_line, column = cols).value
			newws.cell(row =i, column = 1).value = country_name
			newws.cell(row =i, column = 2).value = income_group
			newws.cell(row =i, column = 3).value = year
			newws.cell(row =i, column = 4).value = data_point_2
			newws.cell(row =i, column = 5).value = data_point
			i=i+1



newwb.save(str(out_file))

#employment_agriculture is done